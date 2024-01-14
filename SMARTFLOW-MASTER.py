from tkinter import *
from openpyxl import Workbook
import openpyxl
import zmq
import threading
import datetime
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl import load_workbook
from PIL import Image
from openpyxl.drawing.image import Image
import time
import requests
from bs4 import BeautifulSoup
import os

def Atualizar_bloco():
    try:
        arquivo = open(string_caminho+f'Bloco{numero}.txt','r')
        cont = arquivo.read()
        texto_string = ''.join(cont)
        aba_texto.delete(1.0,END)
        aba_texto.insert(END,texto_string)
        nome_bloco = str(pegar_nome_bloco())
        nome_bloco_atual.config(text=nome_bloco)
    except:
        retorno_dialog.config(text=f'Não existe Bloco{numero}.txt')
'''
def comunication_recebe():
    global status_online
    context = zmq.Context()
    rep_socket = context.socket(zmq.REP)
    rep_socket.bind("tcp://127.0.0.1:5555")

    while(True):
        request = rep_socket.recv_string()
        rep_socket.send_string('teste') 
        status_online = request     
        if status_online != '':
            online.config(text='Online')
        else:
            online.config(text='Offline')
'''
def pegar_nome_bloco():
        arquivo = open(string_caminho+f'Bloco{numero}.txt','r')
        nome_bloco = arquivo.readlines(1)
        nome_bloco_sem_nova_linha = [linha.rstrip() for linha in nome_bloco]  #REMOVE O \N DO FINAL
        nome_bloco_string = ''.join(nome_bloco_sem_nova_linha) #REMOVE ['NOME'] TRANSFORMA LISTA EM STRING
        return nome_bloco_string

def apagar_conteudo_textoaba():
    aba_texto.delete(1.0,END)
    nome_bloco_atual.config(text='')

def apagar_dialog():
        retorno_dialog.config(text='')
      
def apagar_conteudo_criar_planilha():
    criar_planilhas.delete(1.0,END)

def avancar_blocos_botao():
        global numero
        numero += 1

        try:
                prox_bloco = open(string_caminho+f'Bloco{numero}.txt','r')
                arquivo_prox_bloco = prox_bloco.read()
                string_arquivo_prox_bloco = ''.join(arquivo_prox_bloco)
                aba_texto.delete(1.0,END)
                aba_texto.insert(END,string_arquivo_prox_bloco)
                nome_bloco = str(pegar_nome_bloco())
                nome_bloco_atual.config(text=nome_bloco)
        except:
              retorno_dialog.config(text='Não há mais blocos.')

def retroceder_blocos_botao():
        global numero
        numero = numero
        numero -= 1

        if numero < 0:
              numero = 0

        try:
                prox_bloco = open(string_caminho+f'Bloco{numero}.txt','r')
                arquivo_prox_bloco = prox_bloco.read()
                string_arquivo_prox_bloco = ''.join(arquivo_prox_bloco)
                aba_texto.delete(1.0,END)
                aba_texto.insert(END,string_arquivo_prox_bloco)
                nome_bloco = str(pegar_nome_bloco())
                nome_bloco_atual.config(text=nome_bloco)
        except:
              retorno_dialog.config(text='Não há mais blocos.')

def criar_planilhas_novas():

    mes_automatico_atual = datetime.datetime.now()
    mes_automatico_atual.month

    try:  
        if mes_automatico_atual.month == 1:
            nome_mes_planilha = 'JANEIRO'
        elif mes_automatico_atual.month == 2:
                nome_mes_planilha = 'FEVEREIRO'
        elif mes_automatico_atual.month == 3:
                nome_mes_planilha = 'MARÇO'
        elif mes_automatico_atual.month == 4:
                nome_mes_planilha = 'ABRIL'
        elif mes_automatico_atual.month == 5:
                nome_mes_planilha = 'MAIO'
        elif mes_automatico_atual.month == 6:
                nome_mes_planilha = 'JUNHO'
        elif mes_automatico_atual.month == 7:
                nome_mes_planilha = 'JULHO'
        elif mes_automatico_atual.month == 8:
                nome_mes_planilha = 'AGOSTO'
        elif mes_automatico_atual.month == 9:
                nome_mes_planilha = 'SETEMBRO'
        elif mes_automatico_atual.month == 10:
                nome_mes_planilha = 'OUTUBRO'
        elif mes_automatico_atual.month == 11:
                nome_mes_planilha = 'NOVEMBRO'
        elif mes_automatico_atual.month == 12:
                nome_mes_planilha = 'DEZEMBRO'
        else:
            nome_mes_planilha = 'Mes_inexistente'  
       
        nome_bloco = criar_planilhas.get(1.0, END+"-1c") #PEGAR O NOME DE CADA ARQUIVO SEM O /N DO FINAL

        if nome_bloco == '':
           retorno_dialog.config(text='Digite o nome da planilha!')
           return 1

        bloco_separado = nome_bloco.split('\n')
        arq_plani = [[palavra] for palavra in bloco_separado]

        for i in range(len(arq_plani)):
            salvar_arquivo = ''.join(arq_plani[i])

            wb = load_workbook('modelo.xlsx')
            ws = wb.active
            ws = wb['Sheet']

            img = Image('tvco_logo.png')
            img.height = 85 
            img.width = 95
            img2 = Image('sbt_logo.png')
            img2.height = 85 
            img2.width = 90

            ws.add_image(img,'A3')
            ws.add_image(img2,'B3')

            ws['K5'] = salvar_arquivo
            ws['K7'] = nome_mes_planilha

            wb.save(local_planilhas+salvar_arquivo+'.xlsx')

            retorno_dialog.config(text='Criado com sucesso!')

    except Exception as e:
            print(e)
            retorno_dialog.config(text='Por favor, Preencha o campo corretamente.')

def Aplicar_horarios():
      
        horario_selecionado = horario_anotar.get()
        lista_comerciais = str(aba_texto.get(1.0,END+"-1c"))

        if horario_anotar.get() == '':
                retorno_dialog.config(text='Campo horário vazio') 
                return 1 
        elif lista_comerciais == '':
                retorno_dialog.config(text='Campo grade vazio')
                return 1
             
        '''
        file = open(string_caminho+f'Bloco{numero}.txt','r')
        contador = 0
        txt_bloco = file.read()
        conteudo = txt_bloco.split('\n')
        
        for i in conteudo:
              if i:
                    contador += 1
        print('quantidade de linhas: ',contador)
        '''

        file2 = open(string_caminho+f'Bloco{numero}.txt','a')
        content_horario = horario_anotar.get()
        file2.write('\n'+'RODOU - '+content_horario)
        file2.close()

        bloco_separado = lista_comerciais.split('\n')
        comerciais_listas = [[palavra] for palavra in bloco_separado]

        dado_dia = int(dia_atual.get())
        dia_mes = dado_dia + 10

        contador = 0
  
        for i in range(len(comerciais_listas)):          
                try:
                        if contador < (len(comerciais_listas)- 1):  

                                salvar_arquivo = ''.join(comerciais_listas[i + 1])
                                try:
                                        wb = load_workbook(local_planilhas+salvar_arquivo+'.xlsx')                        
                                        ws = wb.active
                                        ws = wb['Sheet']      
                                except:
                                       retorno_dialog.config(text=f'{salvar_arquivo} Não existe na pasta')
                                       return 1                                                               
                                coluna = 0
                                for x in range(16):
                                        coluna += 1
                                        if ws.cell(dia_mes,coluna).value is None:                                               
                                                ws.cell(dia_mes,coluna, value=horario_selecionado)
                                                wb.save(local_planilhas+salvar_arquivo+'.xlsx')
                                                contador += 1
                                                retorno_dialog.config(text='Horário adicionado.')                                     
                                                break
                                        else:
                                                retorno_dialog.config(text=f'Não foi possivel adicionar --> {salvar_arquivo}')
                        else:
                                Atualizar_bloco()

                except Exception as e:
                        retorno_dialog.config(text='Verifique os dados')
                        
def verificar_chave():
       chave_digitada = str(cont_chave.get(1.0, END+'-1c'))
       chave_sistema =  str(sheet_key_nuvem())

       if chave_digitada == chave_sistema:
              chave_certa_incorreta.config(text='Chave correta!')
              arquivo_chave = open('key.txt','w')
              arquivo_chave.writelines(chave_digitada)
              arquivo_chave.close()
              chave_ativacao.destroy()
       else:
              chave_certa_incorreta.config(text='Chave incorreta')

def salvar_caminho():
     arquivo_caminho = open('loc.txt','w')
     location = caminho_grades.get(1.0,END+'-1c')
     arquivo_caminho.writelines(location)
     arquivo_caminho.close

def salvar_caminho_grades():
     arquivo_caminho = open('loc_plan.txt','w')
     location = caminho_grades.get(1.0,END+'-1c')
     arquivo_caminho.writelines(location)
     arquivo_caminho.close
    
def sheet_key_nuvem():     
        url = 'https://progtechsvg.github.io/generator/'
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        paragrafo = soup.find('p').text
        return(paragrafo)
        
if __name__=='__main__':

    url = 'https://progtechsvg.github.io/generator/'
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    version_app = soup.find('h2').text
    desenvolvedor_text = soup.find('h1').text
    
    caminho_salvo = open('loc.txt','r')
    caminho_loc = caminho_salvo.readlines()
    string_caminho = ''.join(caminho_loc)
    caminho_salvo.close()

    caminho_loc_plan = open('loc_plan.txt','r')
    caminho_loc_p = caminho_loc_plan.readlines()
    local_planilhas = ''.join(caminho_loc_p)
    caminho_loc_plan.close()

    status_online = str(version_app)
    day_hoje = datetime.datetime.now()
    horario_atual = datetime.datetime.now()
    numero = 0
    nome_bloco = str(pegar_nome_bloco())

    font_padrao ='Tahoma'
    cor_padrao = '#2E8B57'
    cor_padrao_caixas = 'white'
    cor_padrao_fontes = 'White'
    chave_sist = str(sheet_key_nuvem())

    arquivo_txt_key = open('key.txt', 'r')
    key_verific = arquivo_txt_key.readlines()
    chave_ativ_cliente = ''.join(key_verific)


    if chave_ativ_cliente == '':

        chave_ativacao = Tk()
        chave_ativacao.title('Autenticação')
        chave_ativacao.geometry('500x250')
        chave_ativacao.configure(background=cor_padrao)
        photo = PhotoImage(file='icon.png')
        chave_ativacao.iconphoto(False,photo)
        chave_ativacao.resizable(0,0)

        insira_chave = Label(chave_ativacao,text='Insira sua chave',bg=cor_padrao,fg='white')
        insira_chave.place(x=120,y=40,width=240,height=40)
        insira_chave.config(font=(font_padrao,18))

        cont_chave = Text(chave_ativacao)
        cont_chave.place(x=120, y=80, width=240, height=40)
        cont_chave.config(font=(font_padrao,14))

        Confirmar_botao = Button(chave_ativacao,command=verificar_chave,text='Confirmar',bg='#dde',fg='Black')
        Confirmar_botao.place(x=185,y=130,width=100,height=40)
        Confirmar_botao.config(font=(font_padrao,12))

        chave_certa_incorreta = Label(chave_ativacao,text='',bg=cor_padrao,fg='white')
        chave_certa_incorreta.place(x=140,y=190,width=200,height=40)
        chave_certa_incorreta.config(font=(font_padrao,14)) 
        chave_ativacao.mainloop()

    elif chave_ativ_cliente == chave_sist:
                      
        janela = Tk()
        janela.title('SMART FLOW - MASTER')
        janela.geometry('1280x720')
        janela.configure(background=cor_padrao)
        photo = PhotoImage(file='icon.png')
        janela.iconphoto(False,photo)
        janela.resizable(0,0)

        execute = True

        while(execute):

                #t = threading.Thread(target=comunication_recebe)
                #t.start()

                botao_salvar_caminho = Button(janela,command=salvar_caminho, text='Salvar', bg='#dde', fg='Black')
                botao_salvar_caminho.place(x=880,y=310,width=100,height=25)
                botao_salvar_caminho.config(font=(font_padrao,12))

                botao_salvar_caminho_grades = Button(janela,command=salvar_caminho_grades, text='Salvar', bg='#dde', fg='Black')
                botao_salvar_caminho_grades.place(x=880,y=280,width=100,height=25)
                botao_salvar_caminho_grades.config(font=(font_padrao,12))

                botton_atualizar = Button(janela,command=Atualizar_bloco,text='Atualizar',bg='#dde', fg='Black')
                botton_atualizar.place(x=45, y=640, width=170, height=40)
                botton_atualizar.config(font=(font_padrao,12))

                botton_apagar = Button(janela,command=apagar_conteudo_textoaba,text='Limpar',bg='#dde', fg='Black')
                botton_apagar.place(x=220, y=640, width=150, height=40)
                botton_apagar.config(font=(font_padrao,12))

                botton_apagar2 = Button(janela,command=apagar_conteudo_criar_planilha,text='Limpar',bg='#dde', fg='Black')
                botton_apagar2.place(x=730, y=560, width=150, height=40)
                botton_apagar2.config(font=(font_padrao,12))

                botton_adiantar = Button(janela,command=avancar_blocos_botao,text='>',bg='#dde', fg='Black')
                botton_adiantar.place(x=460, y=640, width=85, height=40)
                botton_adiantar.config(font=(font_padrao,12))

                botton_retroceder = Button(janela,command=retroceder_blocos_botao,text='<',bg='#dde', fg='Black')
                botton_retroceder.place(x=375, y=640, width=85, height=40)
                botton_retroceder.config(font=(font_padrao,12))

                botao_aplicar = Button(janela,command=Aplicar_horarios,text='Aplicar',bg='#dde', fg='Black')
                botao_aplicar.place(x=1160, y=640, width=85, height=40)
                botao_aplicar.config(font=(font_padrao,12))

                botao_limpar_dialog = Button(janela,command=apagar_dialog,text='Limpar',bg='#dde', fg='Black')
                botao_limpar_dialog.place(x=1070, y=640, width=85, height=40)
                botao_limpar_dialog.config(font=(font_padrao,12))

                botao_criar_planilhas= Button(janela,command=criar_planilhas_novas,text='Criar Planilha',bg='#dde', fg='Black')
                botao_criar_planilhas.place(x=575, y=560, width=150, height=40)
                botao_criar_planilhas.config(font=(font_padrao,12))

                criar_planilhas = Text(janela,bg=cor_padrao_caixas)
                criar_planilhas.place(x=575, y=350, width=670, height=200)
                Scrollbar2 = Scrollbar(janela, command=criar_planilhas.yview)
                Scrollbar2.place(x=1225,y=350,width=20,height=200)
                criar_planilhas.config(yscrollcommand=Scrollbar2.set,bg=cor_padrao_caixas,font=(font_padrao,13))

                caminho_grades = Text(janela)
                caminho_grades.place(x=575, y=310,width=300, height=25)
                caminho_grades.config(font=(font_padrao,12))
                caminho_grades.insert(END,string_caminho)

                caminho_planilhas = Text(janela)
                caminho_planilhas.place(x=575, y=280,width=300, height=25)
                caminho_planilhas.config(font=(font_padrao,12))
                caminho_planilhas.insert(END,local_planilhas)

                aba_texto = Text(janela)
                aba_texto.place(x=45, y=30, width=500,height=600)
                scrollbar = Scrollbar(janela, command=aba_texto.yview)
                scrollbar.place(x=525, y=30, width=20,height=600)
                aba_texto.config(yscrollcommand=scrollbar.set,bg=cor_padrao_caixas,font=(font_padrao,14))

                retorno_dialog = Label(janela, bg=cor_padrao_caixas, fg='Black')
                retorno_dialog.place(x=575, y=640, width=480, height=40)
                retorno_dialog.config(font=(font_padrao,14))

                horario_anotar = Entry(janela)
                #hour_atual = (str(horario_atual.hour)+':'+str(horario_atual.minute))
                #horario_anotar.insert(10,hour_atual)
                horario_anotar.place(x=740, y=120, width=120, height=40)

                label_digite_horario = Label(janela, text='Horário:',font=font_padrao,bg=cor_padrao, fg=cor_padrao_fontes)
                label_digite_horario.place(x=550, y=120, width=150, height=40)
                label_digite_horario.config(font=(font_padrao,14))

                online = Label(janela,text=status_online,font=font_padrao,bg=cor_padrao , fg=cor_padrao_fontes)
                online.place(x=1185, y=10, width=100, height=30)
                online.config(font=(font_padrao,12))

                desenvolvedor = Label(janela,text=desenvolvedor_text,font=font_padrao,bg=cor_padrao , fg=cor_padrao_fontes)
                desenvolvedor.place(x=1185, y=35, width=100, height=30)
                desenvolvedor.config(font=(font_padrao,10))

                bloco_atual_indicador = Label(janela,text='Bloco Atual:',bg=cor_padrao, fg=cor_padrao_fontes)
                bloco_atual_indicador.place(x=550, y=40,width=150, height=30)
                bloco_atual_indicador.config(font=(font_padrao,14))

                nome_bloco_atual = Label(janela,bg=cor_padrao, fg=cor_padrao_fontes)
                nome_bloco_atual.place(x=700, y=40,width=400, height=30)
                nome_bloco_atual.config(font=(font_padrao,14),text=nome_bloco)

                indicar_dia_atual = Label(janela,text='Dia atual:',bg=cor_padrao, fg=cor_padrao_fontes)
                indicar_dia_atual.place(x=550, y=80,width=150, height=30)
                indicar_dia_atual.config(font=(font_padrao,14))

                dia_atual = Entry(janela, font=font_padrao, fg='Black')
                dia_atual.insert(5,day_hoje.day)
                dia_atual.place(x=740, y=80,width=120, height=30)
                        
                Atualizar_bloco()
                janela.mainloop()
                execute = False
    else:
        arquivo_txt_key = open('key.txt', 'w')
        pass